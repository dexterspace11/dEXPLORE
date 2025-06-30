# ---------------- Enhanced CNN-EQIC Cluster Analysis App with Interpretive Insights ----------------

# Imports
import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from sklearn.impute import SimpleImputer
from sklearn.preprocessing import MinMaxScaler
from sklearn.decomposition import PCA
from sklearn.metrics import silhouette_score, davies_bouldin_score, calinski_harabasz_score
from sklearn.cluster import KMeans, DBSCAN
from sklearn.linear_model import LinearRegression
from datetime import datetime
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment

# ---------------- Memory Structures ----------------
class EpisodicMemory:
    def __init__(self):
        self.episodes = {}
        self.current_episode = None

    def create_episode(self, timestamp):
        self.current_episode = timestamp
        self.episodes[timestamp] = {'patterns': [], 'emotional_tags': []}

    def store_pattern(self, pattern, emotional_tag):
        if self.current_episode is None:
            self.create_episode(datetime.now())
        self.episodes[self.current_episode]['patterns'].append(pattern)
        self.episodes[self.current_episode]['emotional_tags'].append(emotional_tag)

class WorkingMemory:
    def __init__(self, capacity=20):
        self.capacity = capacity
        self.short_term_patterns = []
        self.temporal_context = []

    def store(self, pattern, timestamp):
        if len(self.short_term_patterns) >= self.capacity:
            self.short_term_patterns.pop(0)
            self.temporal_context.pop(0)
        self.short_term_patterns.append(pattern)
        self.temporal_context.append(timestamp)

# ---------------- Neural Unit ----------------
class HybridNeuralUnit:
    def __init__(self, position, decay_rate=100.0):
        self.position = position
        self.age = 0
        self.usage_count = 0
        self.last_spike_time = None
        self.decay_rate = decay_rate
        self.connections = []

    def distance(self, input_pattern):
        diff = np.abs(input_pattern - self.position)
        dist = np.sqrt(np.sum(diff ** 2))
        decay = np.exp(-self.age / self.decay_rate)
        return (np.exp(-2.0 * dist) + 0.5 / (1 + 0.9 * dist)) * decay

    def update_spike_time(self):
        self.last_spike_time = datetime.now()

    def communicate(self, other_units):
        for other in other_units:
            if other != self:
                dist = np.linalg.norm(self.position - other.position)
                if dist < 0.5 and other not in self.connections:
                    self.connections.append(other)

# ---------------- Neural Network ----------------
class HybridNeuralNetwork:
    def __init__(self, working_memory_capacity=20, decay_rate=100.0):
        self.units = []
        self.episodic_memory = EpisodicMemory()
        self.working_memory = WorkingMemory(capacity=working_memory_capacity)
        self.decay_rate = decay_rate

    def generate_unit(self, position):
        unit = HybridNeuralUnit(position, decay_rate=self.decay_rate)
        unit.communicate(self.units)
        self.units.append(unit)
        return unit

    def process_input(self, input_pattern):
        if not self.units:
            return self.generate_unit(input_pattern), 0.0

        similarities = [(unit, unit.distance(input_pattern)) for unit in self.units]
        similarities.sort(key=lambda x: x[1], reverse=True)
        best_unit, best_similarity = similarities[0]

        self.episodic_memory.store_pattern(input_pattern, best_similarity)
        self.working_memory.store(input_pattern, datetime.now())
        best_unit.age = 0
        best_unit.usage_count += 1
        best_unit.update_spike_time()

        if best_similarity < 0.6:
            return self.generate_unit(input_pattern), best_similarity

        return best_unit, best_similarity

# ---------------- Forecasting ----------------
def forecast_next(values, steps=1):
    model = LinearRegression()
    X = np.arange(len(values)).reshape(-1, 1)
    y = values
    model.fit(X, y)
    X_pred = np.arange(len(values), len(values) + steps).reshape(-1, 1)
    return model.predict(X_pred)

# ---------------- Clustering Evaluation ----------------
def evaluate_clustering(X, labels):
    try:
        if len(np.unique(labels)) > 1 and len(X) > len(np.unique(labels)):
            sil = silhouette_score(X, labels)
            db = davies_bouldin_score(X, labels)
            ch = calinski_harabasz_score(X, labels)
            return {'Silhouette': sil, 'Davies-Bouldin': db, 'Calinski-Harabasz': ch}
        else:
            return {'Silhouette': np.nan, 'Davies-Bouldin': np.nan, 'Calinski-Harabasz': np.nan}
    except:
        return {'Silhouette': np.nan, 'Davies-Bouldin': np.nan, 'Calinski-Harabasz': np.nan}

# ---------------- Excel Export ----------------
def save_to_excel(path, df_original, selected_features, patterns_df,
                  cnn_labels, kmeans_labels, dbscan_labels,
                  cnn_metrics, kmeans_metrics, dbscan_metrics,
                  cluster_summaries, cluster_corrs, narrative_report):

    wb = openpyxl.Workbook()

    # Sheet1: Raw Data (selected features)
    ws1 = wb.active
    ws1.title = "Selected Features"
    for r in dataframe_to_rows(df_original[selected_features], index=False, header=True):
        ws1.append(r)
    for cell in ws1[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    ws1.auto_filter.ref = ws1.dimensions
    for col_cells in ws1.columns:
        length = max(len(str(cell.value)) for cell in col_cells)
        ws1.column_dimensions[col_cells[0].column_letter].width = length + 2

    # Sheet2: Cluster Labels Summary
    ws2 = wb.create_sheet("Cluster Summary")
    summary_data = {
        'Algorithm': ['CNN-EQIC', 'KMeans', 'DBSCAN'],
        'Clusters': [len(np.unique(cnn_labels)), len(np.unique(kmeans_labels)), len(np.unique(dbscan_labels)) - (1 if -1 in dbscan_labels else 0)],
        'Outliers': [0, 0, sum(dbscan_labels == -1)],
        'Silhouette': [cnn_metrics['Silhouette'], kmeans_metrics['Silhouette'], dbscan_metrics['Silhouette']],
        'Davies-Bouldin': [cnn_metrics['Davies-Bouldin'], kmeans_metrics['Davies-Bouldin'], dbscan_metrics['Davies-Bouldin']],
        'Calinski-Harabasz': [cnn_metrics['Calinski-Harabasz'], kmeans_metrics['Calinski-Harabasz'], dbscan_metrics['Calinski-Harabasz']]
    }
    summary_df = pd.DataFrame(summary_data)
    for r in dataframe_to_rows(summary_df, index=False, header=True):
        ws2.append(r)
    for cell in ws2[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    ws2.auto_filter.ref = ws2.dimensions
    for col_cells in ws2.columns:
        length = max(len(str(cell.value)) for cell in col_cells)
        ws2.column_dimensions[col_cells[0].column_letter].width = length + 2

    # Sheet3: Feature Importance by Cluster (Table format)
    ws3 = wb.create_sheet("Feature Importance")

    # CNN-EQIC Feature Importance
    ws3.append(["CNN-EQIC Feature Importance"])
    for label, dfc in cluster_summaries['CNN'].items():
        ws3.append([f"Unit {label} Mean Feature Values"])
        fi_df = dfc.round(4)
        ws3.append(list(fi_df.columns))
        ws3.append(list(fi_df.values[0]))
        ws3.append([])

    # KMeans Feature Importance
    ws3.append(["KMeans Feature Importance"])
    for label, dfc in cluster_summaries['KMeans'].items():
        ws3.append([f"Cluster {label} Mean Feature Values"])
        fi_df = dfc.round(4)
        ws3.append(list(fi_df.columns))
        ws3.append(list(fi_df.values[0]))
        ws3.append([])

    # DBSCAN Feature Importance (ignore noise -1)
    ws3.append(["DBSCAN Feature Importance"])
    for label, dfc in cluster_summaries['DBSCAN'].items():
        ws3.append([f"Cluster {label} Mean Feature Values"])
        fi_df = dfc.round(4)
        ws3.append(list(fi_df.columns))
        ws3.append(list(fi_df.values[0]))
        ws3.append([])

    # Sheet4+: Cluster-wise Correlations
    for method in cluster_corrs.keys():
        ws = wb.create_sheet(f"{method} Corr")
        ws.append([f"{method} Cluster-wise Correlations"])
        for label, corr_df in cluster_corrs[method].items():
            ws.append([f"Cluster {label} Correlation Matrix"])
            for r in dataframe_to_rows(corr_df.round(4), index=True, header=True):
                ws.append(r)
            ws.append([])

    # Sheet last: Narrative Report
    wsn = wb.create_sheet("Narrative Report")
    for line in narrative_report:
        wsn.append([line])
    wsn.column_dimensions['A'].width = 120
    wsn["A1"].font = Font(bold=True, size=14)

    wb.save(path)

# ... [Keep all your imports and class definitions as-is] ...

# ---------------- Streamlit UI ----------------
st.set_page_config(page_title="Enhanced CNN-EQIC Cluster App", layout="wide")
st.title("🔍 Hybrid CNN-EQIC Clustering Explorer")

uploaded_file = st.file_uploader("Upload a CSV or Excel file", type=["csv", "xlsx"])

if uploaded_file:
    df = pd.read_csv(uploaded_file) if uploaded_file.name.endswith(".csv") else pd.read_excel(uploaded_file)
    st.markdown("### Data Preview")
    st.dataframe(df.head())

    numerical_cols = df.select_dtypes(include=[np.number]).columns.tolist()
    selected = st.multiselect("Select features", numerical_cols, default=numerical_cols[:4])
    window_size = st.slider("Window Size", 2, 20, 5)

    # NEW: User inputs for clustering parameters
    n_clusters = st.number_input("Number of Clusters for KMeans", min_value=2, max_value=20, value=5, step=1)
    dbscan_eps = st.slider("DBSCAN Epsilon (Neighborhood Radius)", 0.1, 1.0, 0.4, 0.05)

    clean = SimpleImputer().fit_transform(df[selected])
    scaled = MinMaxScaler().fit_transform(clean)

    param_grid = [{'working_memory_capacity': c, 'decay_rate': d} for c in [10, 20] for d in [50.0, 100.0]]
    def evaluate(net, data):
        scores = []
        for i in range(window_size, len(data)):
            pattern = data[i - window_size:i].flatten()
            _, sim = net.process_input(pattern)
            scores.append(sim)
        return np.mean(scores)

    best_score = -np.inf
    best_params = {}
    for params in param_grid:
        net = HybridNeuralNetwork(**params)
        score = evaluate(net, scaled)
        if score > best_score:
            best_params, best_score = params, score

    net = HybridNeuralNetwork(**best_params)
    similarities, timestamps = [], []
    for i in range(window_size, len(scaled)):
        pattern = scaled[i - window_size:i].flatten()
        _, sim = net.process_input(pattern)
        similarities.append(sim)
        timestamps.append(i)

    st.success(f"Best Params: {best_params}, Best Similarity: {best_score:.4f}")

    patterns = [p for e in net.episodic_memory.episodes.values() for p in e['patterns']]
    if patterns:
        patterns = np.array(patterns)
        pcs = PCA(n_components=2).fit_transform(patterns)

        # CNN Labels (assign pattern to unit index repeated by usage_count, trim or pad to patterns length)
        cnn_labels = []
        for i, unit in enumerate(net.units):
            cnn_labels.extend([i] * unit.usage_count)
        cnn_labels = cnn_labels[:len(patterns)]

        # Evaluate clustering metrics on PCA projection
        cnn_metrics = evaluate_clustering(pcs, cnn_labels)
        kmeans = KMeans(n_clusters=n_clusters, random_state=42).fit(patterns)
        kmeans_metrics = evaluate_clustering(pcs, kmeans.labels_)
        db = DBSCAN(eps=dbscan_eps, min_samples=5).fit(patterns)
        dbscan_metrics = evaluate_clustering(pcs, db.labels_)

        # Summary dataframe for display
        summary_df = pd.DataFrame({
            'Algorithm': ['CNN-EQIC', 'KMeans', 'DBSCAN'],
            'Clusters': [len(np.unique(cnn_labels)),
                         len(np.unique(kmeans.labels_)),
                         len(np.unique(db.labels_)) - (1 if -1 in db.labels_ else 0)],
            'Outliers': [0, 0, sum(db.labels_ == -1)],
            'Silhouette': [cnn_metrics['Silhouette'], kmeans_metrics['Silhouette'], dbscan_metrics['Silhouette']],
            'Davies-Bouldin': [cnn_metrics['Davies-Bouldin'], kmeans_metrics['Davies-Bouldin'], dbscan_metrics['Davies-Bouldin']],
            'Calinski-Harabasz': [cnn_metrics['Calinski-Harabasz'], kmeans_metrics['Calinski-Harabasz'], dbscan_metrics['Calinski-Harabasz']]
        })
        st.markdown("### Neural Cluster Summary & Metrics")
        st.dataframe(summary_df)

        st.markdown("### PCA Cluster Projections")
        fig, ax = plt.subplots(1, 3, figsize=(18, 5))
        ax[0].scatter(pcs[:, 0], pcs[:, 1], c=cnn_labels, cmap='tab20', s=20)
        ax[0].set_title("CNN-EQIC Clusters")
        ax[1].scatter(pcs[:, 0], pcs[:, 1], c=kmeans.labels_, cmap='Set2', s=20)
        ax[1].set_title(f"KMeans Clusters (k={n_clusters})")
        ax[2].scatter(pcs[:, 0], pcs[:, 1], c=db.labels_, cmap='coolwarm', s=20)
        ax[2].set_title(f"DBSCAN Clusters (eps={dbscan_eps:.2f})")
        st.pyplot(fig)

        # Build cluster summary (mean feature values) and cluster-wise correlation matrices
        cluster_summaries = {'CNN': {}, 'KMeans': {}, 'DBSCAN': {}}
        cluster_corrs = {'CNN': {}, 'KMeans': {}, 'DBSCAN': {}}

        patterns_df = pd.DataFrame(patterns, columns=[f"F{i}" for i in range(patterns.shape[1])])
        # Add cluster labels
        patterns_df['CNN_Label'] = cnn_labels
        patterns_df['KMeans_Label'] = kmeans.labels_
        patterns_df['DBSCAN_Label'] = db.labels_

        # Calculate mean feature values & correlations per cluster
        for method, labels in [('CNN', cnn_labels), ('KMeans', kmeans.labels_), ('DBSCAN', db.labels_)]:
            unique_labels = np.unique(labels)
            # Ignore DBSCAN noise -1 for summaries
            if method == 'DBSCAN':
                unique_labels = unique_labels[unique_labels != -1]
            for label in unique_labels:
                cluster_data = patterns_df[patterns_df[f'{method}_Label'] == label].drop(columns=[col for col in patterns_df.columns if 'Label' in col])
                cluster_summaries[method][label] = cluster_data.mean()
                cluster_corrs[method][label] = cluster_data.corr()

        # Display Feature Importance Tables for CNN clusters (most requested)
        st.markdown("### CNN-EQIC Feature Importance Summary (Mean Feature Values per Cluster)")
        cnn_fi_df = pd.DataFrame(cluster_summaries['CNN']).T
        st.dataframe(cnn_fi_df.style.format("{:.4f}"))

        # Narrative reporting example for clusters
        narrative_report = []
        narrative_report.append("Hybrid CNN-EQIC Clustering Narrative Report\n")
        for method in ['CNN', 'KMeans', 'DBSCAN']:
            narrative_report.append(f"Algorithm: {method}")
            for label, mean_vals in cluster_summaries[method].items():
                narrative_report.append(f" Cluster {label}:")
                narrative_report.append(f"  - Number of points: {len(patterns_df[patterns_df[f'{method}_Label']==label])}")
                narrative_report.append(f"  - Mean feature values:")
                for feat, val in mean_vals.items():
                    narrative_report.append(f"     {feat}: {val:.4f}")
                narrative_report.append("")

        # Export button & logic
        save_path = r"C:\Users\oliva\OneDrive\Documents\Excel doc\CNN_EQIC_Cluster_Analysis.xlsx"
        if st.button("Export Cluster Analysis Results to Excel"):
            save_to_excel(save_path, df, selected, patterns_df,
                          cnn_labels, kmeans.labels_, db.labels_,
                          cnn_metrics, kmeans_metrics, dbscan_metrics,
                          cluster_summaries, cluster_corrs,
                          narrative_report)
            st.success(f"Results exported to {save_path}")

    else:
        st.warning("No patterns found in episodic memory.")

