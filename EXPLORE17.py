# ---------------- Enhanced CNN-EQIC EDA with Professional Cluster Analysis for Research ----------------
import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from sklearn.impute import SimpleImputer
from sklearn.preprocessing import MinMaxScaler
from sklearn.decomposition import PCA
from sklearn.feature_selection import mutual_info_regression
from scipy.stats import zscore
from sklearn.cluster import KMeans, DBSCAN
from sklearn.linear_model import LinearRegression
from sklearn.metrics import silhouette_score, davies_bouldin_score, calinski_harabasz_score
from datetime import datetime
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment
import os

# ---------------- Memory Structures ----------------
class EpisodicMemory:
    def __init__(self):
        self.episodes = {}
        self.current_episode = None

    def create_episode(self, timestamp):
        self.current_episode = timestamp
        self.episodes[timestamp] = {'patterns': [], 'emotional_tags': []}

    def store_pattern(self, pattern, emotional_tag):
        if self.current_episode is None:
            self.create_episode(datetime.now())
        self.episodes[self.current_episode]['patterns'].append(pattern)
        self.episodes[self.current_episode]['emotional_tags'].append(emotional_tag)

class WorkingMemory:
    def __init__(self, capacity=20):
        self.capacity = capacity
        self.short_term_patterns = []
        self.temporal_context = []

    def store(self, pattern, timestamp):
        if len(self.short_term_patterns) >= self.capacity:
            self.short_term_patterns.pop(0)
            self.temporal_context.pop(0)
        self.short_term_patterns.append(pattern)
        self.temporal_context.append(timestamp)

# ---------------- Neural Unit ----------------
class HybridNeuralUnit:
    def __init__(self, position, decay_rate=100.0):
        self.position = position
        self.age = 0
        self.usage_count = 0
        self.last_spike_time = None
        self.decay_rate = decay_rate
        self.connections = []

    def distance(self, input_pattern):
        diff = np.abs(input_pattern - self.position)
        dist = np.sqrt(np.sum(diff ** 2))
        decay = np.exp(-self.age / self.decay_rate)
        return (np.exp(-2.0 * dist) + 0.5 / (1 + 0.9 * dist)) * decay

    def update_spike_time(self):
        self.last_spike_time = datetime.now()

    def communicate(self, other_units):
        for other in other_units:
            if other != self:
                dist = np.linalg.norm(self.position - other.position)
                if dist < 0.5 and other not in self.connections:
                    self.connections.append(other)

# ---------------- Neural Network ----------------
class HybridNeuralNetwork:
    def __init__(self, working_memory_capacity=20, decay_rate=100.0):
        self.units = []
        self.episodic_memory = EpisodicMemory()
        self.working_memory = WorkingMemory(capacity=working_memory_capacity)
        self.decay_rate = decay_rate

    def generate_unit(self, position):
        unit = HybridNeuralUnit(position, decay_rate=self.decay_rate)
        unit.communicate(self.units)
        self.units.append(unit)
        return unit

    def process_input(self, input_pattern):
        if not self.units:
            return self.generate_unit(input_pattern), 0.0

        similarities = [(unit, unit.distance(input_pattern)) for unit in self.units]
        similarities.sort(key=lambda x: x[1], reverse=True)
        best_unit, best_similarity = similarities[0]

        self.episodic_memory.store_pattern(input_pattern, best_similarity)
        self.working_memory.store(input_pattern, datetime.now())
        best_unit.age = 0
        best_unit.usage_count += 1
        best_unit.update_spike_time()

        if best_similarity < 0.6:
            return self.generate_unit(input_pattern), best_similarity

        return best_unit, best_similarity

# ---------------- Forecasting ----------------
def forecast_next(values, steps=1):
    model = LinearRegression()
    X = np.arange(len(values)).reshape(-1, 1)
    y = values
    model.fit(X, y)
    X_pred = np.arange(len(values), len(values) + steps).reshape(-1, 1)
    return model.predict(X_pred)

# ---------------- Excel Export ----------------
def export_detailed_clustering(filepath, best_params, silhouette, db_score, ch_score, summary_text, patterns, kmeans_labels, pcs, centroids):
    wb = openpyxl.Workbook()

    # Summary Sheet
    ws1 = wb.active
    ws1.title = "Clustering Summary"
    rows = [
        ["Metric", "Value"],
        ["Working Memory Capacity", best_params['working_memory_capacity']],
        ["Decay Rate", best_params['decay_rate']],
        ["Silhouette Score", f"{silhouette:.4f}"],
        ["Davies-Bouldin Index", f"{db_score:.4f}"],
        ["Calinski-Harabasz Score", f"{ch_score:.4f}"],
        ["Summary Narrative", summary_text]
    ]
    for row in rows:
        ws1.append(row)
    for cell in ws1[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 100

    # Pattern Assignments
    ws2 = wb.create_sheet("Cluster Assignments")
    df_patterns = pd.DataFrame(patterns)
    df_patterns['KMeans_Cluster'] = kmeans_labels
    for r in dataframe_to_rows(df_patterns, index=False, header=True):
        ws2.append(r)
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    ws2.auto_filter.ref = ws2.dimensions

    # PCA Sheet
    ws3 = wb.create_sheet("PCA Projection")
    pcs_df = pd.DataFrame(pcs, columns=['PC1', 'PC2'])
    pcs_df['Cluster'] = kmeans_labels
    for r in dataframe_to_rows(pcs_df, index=False, header=True):
        ws3.append(r)
    for cell in ws3[1]:
        cell.font = Font(bold=True)
    ws3.auto_filter.ref = ws3.dimensions

    # Centroid Analysis
    ws4 = wb.create_sheet("Centroid Analysis")
    centroid_df = pd.DataFrame(centroids, columns=[f"Feature_{i+1}" for i in range(centroids.shape[1])])
    centroid_df['Cluster'] = range(len(centroids))
    for r in dataframe_to_rows(centroid_df, index=False, header=True):
        ws4.append(r)
    for cell in ws4[1]:
        cell.font = Font(bold=True)
    ws4.auto_filter.ref = ws4.dimensions

    wb.save(filepath)

# ---------------- Streamlit UI ----------------
st.set_page_config(page_title="Advanced CNN-EQIC EDA", layout="wide")
st.title("📊 CNN-EQIC: Research-Ready Cluster Analysis")

uploaded_file = st.file_uploader("Upload a CSV or Excel file", type=["csv", "xlsx"])

if uploaded_file:
    df = pd.read_csv(uploaded_file) if uploaded_file.name.endswith(".csv") else pd.read_excel(uploaded_file)
    st.markdown("### Dataset Preview")
    st.dataframe(df.head())

    numerical_cols = df.select_dtypes(include=[np.number]).columns.tolist()
    selected = st.multiselect("Select Numerical Features for Analysis", numerical_cols, default=numerical_cols[:4])
    window_size = st.slider("Window Size (Pattern Recognition)", 2, 20, 5)

    clean = SimpleImputer().fit_transform(df[selected])
    scaled = MinMaxScaler().fit_transform(clean)

    st.markdown("### Neural Network Training")
    param_grid = [{'working_memory_capacity': c, 'decay_rate': d} for c in [10, 20] for d in [50.0, 100.0]]
    def evaluate(net, data):
        scores = []
        for i in range(window_size, len(data)):
            pattern = data[i - window_size:i].flatten()
            _, sim = net.process_input(pattern)
            scores.append(sim)
        return np.mean(scores)

    best_score, best_params = -np.inf, {}
    for params in param_grid:
        net = HybridNeuralNetwork(**params)
        score = evaluate(net, scaled)
        if score > best_score:
            best_params, best_score = params, score
    st.success(f"Best Network Params: {best_params}, Similarity Score: {best_score:.4f}")

    net = HybridNeuralNetwork(**best_params)
    similarities, timestamps = [], []
    for i in range(window_size, len(scaled)):
        pattern = scaled[i - window_size:i].flatten()
        _, sim = net.process_input(pattern)
        similarities.append(sim)
        timestamps.append(i)

    patterns = [p for e in net.episodic_memory.episodes.values() for p in e['patterns']]
    if patterns:
        patterns = np.array(patterns)
        st.markdown("### KMeans Clustering")
        n_clusters = st.slider("Select K (Clusters)", 2, 10, 5)
        kmeans = KMeans(n_clusters=n_clusters).fit(patterns)
        labels = kmeans.labels_
        centroids = kmeans.cluster_centers_

        silhouette = silhouette_score(patterns, labels)
        db_score = davies_bouldin_score(patterns, labels)
        ch_score = calinski_harabasz_score(patterns, labels)

        summary_text = f"Clusters discovered show {'distinct' if silhouette > 0.5 else 'moderate'} separation. Silhouette Score: {silhouette:.2f}, DB Index: {db_score:.2f}, CH Score: {ch_score:.2f}. Each cluster groups patterns with similar trajectory/structure. Centroid analysis provides the average pattern signature per group."

        st.markdown("#### Clustering Metrics")
        st.code(summary_text)

        pcs = PCA(n_components=2).fit_transform(patterns)
        fig, ax = plt.subplots()
        scatter = ax.scatter(pcs[:, 0], pcs[:, 1], c=labels, cmap='tab10')
        ax.set_title("Cluster Visualization via PCA")
        st.pyplot(fig)

        st.markdown("### Export Results for Research")
        export_path = r"C:\\Users\\oliva\\OneDrive\\Documents\\Excel doc\\DNNanalysis.xlsx"
        if st.button("Export Detailed Excel Report"):
            export_detailed_clustering(export_path, best_params, silhouette, db_score, ch_score, summary_text, patterns, labels, pcs, centroids)
            st.success(f"📁 Results saved to {export_path}")

        st.markdown("### Interpretive Notes")
        st.info("""
        - Each cluster represents a group of time-based patterns with shared statistical shape.
        - Centroids indicate average pattern of each group, useful for behavioral interpretation.
        - Outliers may indicate rare, novel, or transitional states.
        - PCA shows pattern group overlap or distinctness.
        - These clusters can be mapped to real-world segments, regimes, or transitions.
        """)
    else:
        st.warning("Insufficient patterns detected. Adjust window size or check data range.")