# ------------------- Full Updated CNN-EQIC Cluster-Focused Streamlit App -------------------
import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from sklearn.impute import SimpleImputer
from sklearn.preprocessing import MinMaxScaler
from sklearn.decomposition import PCA
from sklearn.cluster import KMeans, DBSCAN
from datetime import datetime
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment

# ---------------- Memory Structures ----------------
class EpisodicMemory:
    def __init__(self):
        self.episodes = {}
        self.current_episode = None

    def create_episode(self, timestamp):
        self.current_episode = timestamp
        self.episodes[timestamp] = {'patterns': [], 'emotional_tags': []}

    def store_pattern(self, pattern, emotional_tag):
        if self.current_episode is None:
            self.create_episode(datetime.now())
        self.episodes[self.current_episode]['patterns'].append(pattern)
        self.episodes[self.current_episode]['emotional_tags'].append(emotional_tag)

class WorkingMemory:
    def __init__(self, capacity=20):
        self.capacity = capacity
        self.short_term_patterns = []
        self.temporal_context = []

    def store(self, pattern, timestamp):
        if len(self.short_term_patterns) >= self.capacity:
            self.short_term_patterns.pop(0)
            self.temporal_context.pop(0)
        self.short_term_patterns.append(pattern)
        self.temporal_context.append(timestamp)

# ---------------- Neural Unit ----------------
class HybridNeuralUnit:
    def __init__(self, position, decay_rate=100.0):
        self.position = position
        self.age = 0
        self.usage_count = 0
        self.last_spike_time = None
        self.decay_rate = decay_rate
        self.connections = []

    def distance(self, input_pattern):
        diff = np.abs(input_pattern - self.position)
        dist = np.sqrt(np.sum(diff ** 2))
        decay = np.exp(-self.age / self.decay_rate)
        return (np.exp(-2.0 * dist) + 0.5 / (1 + 0.9 * dist)) * decay

    def update_spike_time(self):
        self.last_spike_time = datetime.now()

    def communicate(self, other_units):
        for other in other_units:
            if other != self:
                dist = np.linalg.norm(self.position - other.position)
                if dist < 0.5 and other not in self.connections:
                    self.connections.append(other)

# ---------------- Neural Network ----------------
class HybridNeuralNetwork:
    def __init__(self, working_memory_capacity=20, decay_rate=100.0):
        self.units = []
        self.episodic_memory = EpisodicMemory()
        self.working_memory = WorkingMemory(capacity=working_memory_capacity)
        self.decay_rate = decay_rate

    def generate_unit(self, position):
        unit = HybridNeuralUnit(position, decay_rate=self.decay_rate)
        unit.communicate(self.units)
        self.units.append(unit)
        return unit

    def process_input(self, input_pattern):
        if not self.units:
            return self.generate_unit(input_pattern), 0.0

        similarities = [(unit, unit.distance(input_pattern)) for unit in self.units]
        similarities.sort(key=lambda x: x[1], reverse=True)
        best_unit, best_similarity = similarities[0]

        self.episodic_memory.store_pattern(input_pattern, best_similarity)
        self.working_memory.store(input_pattern, datetime.now())
        best_unit.age = 0
        best_unit.usage_count += 1
        best_unit.update_spike_time()

        if best_similarity < 0.6:
            return self.generate_unit(input_pattern), best_similarity

        return best_unit, best_similarity

# ---------------- Excel Export ----------------
def save_to_excel(save_path, df, selected, cnn_labels, kmeans_labels, dbscan_labels, pcs):
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Clean Data"
    for r in dataframe_to_rows(df[selected], index=False, header=True):
        ws1.append(r)
    for cell in ws1[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    ws2 = wb.create_sheet("Clustering Labels")
    cluster_df = pd.DataFrame({
        "Pattern Index": range(len(cnn_labels)),
        "CNN-EQIC Cluster": cnn_labels,
        "KMeans Cluster": kmeans_labels,
        "DBSCAN Cluster": dbscan_labels
    })
    for r in dataframe_to_rows(cluster_df, index=False, header=True):
        ws2.append(r)

    ws3 = wb.create_sheet("PCA Projection")
    pcs_df = pd.DataFrame(pcs, columns=['PC1', 'PC2'])
    for r in dataframe_to_rows(pcs_df, index=False, header=True):
        ws3.append(r)

    ws4 = wb.create_sheet("Summary")
    ws4.append(["Total Patterns", len(cnn_labels)])
    ws4.append(["CNN-EQIC Clusters", len(set(cnn_labels))])
    ws4.append(["KMeans Clusters", len(set(kmeans_labels))])
    ws4.append(["DBSCAN Clusters (non-outlier)", len(set(dbscan_labels)) - (1 if -1 in dbscan_labels else 0)])
    ws4.append(["DBSCAN Outliers", sum(np.array(dbscan_labels) == -1)])

    wb.save(save_path)

# ---------------- Streamlit UI ----------------
st.set_page_config(page_title="CNN-EQIC Clustering Focus", layout="wide")
st.title("📊 CNN-EQIC Cluster Analysis vs KMeans/DBSCAN")

uploaded_file = st.file_uploader("Upload CSV or Excel", type=["csv", "xlsx"])

if uploaded_file:
    df = pd.read_csv(uploaded_file) if uploaded_file.name.endswith(".csv") else pd.read_excel(uploaded_file)
    st.write("### Data Preview")
    st.dataframe(df.head())

    num_cols = df.select_dtypes(include=[np.number]).columns.tolist()
    selected = st.multiselect("Select Features", num_cols, default=num_cols[:4])
    window_size = st.slider("Window Size", 2, 20, 5)

    clean = SimpleImputer().fit_transform(df[selected])
    scaled = MinMaxScaler().fit_transform(clean)

    param_grid = [{'working_memory_capacity': c, 'decay_rate': d} for c in [10, 20] for d in [50.0, 100.0]]

    def evaluate(net, data):
        scores = []
        for i in range(window_size, len(data)):
            pattern = data[i - window_size:i].flatten()
            _, sim = net.process_input(pattern)
            scores.append(sim)
        return np.mean(scores)

    best_score = -np.inf
    best_params = {}
    for params in param_grid:
        net = HybridNeuralNetwork(**params)
        score = evaluate(net, scaled)
        if score > best_score:
            best_params, best_score = params, score

    st.success(f"Best Params: {best_params}, Similarity Score: {best_score:.4f}")

    net = HybridNeuralNetwork(**best_params)
    for i in range(window_size, len(scaled)):
        pattern = scaled[i - window_size:i].flatten()
        net.process_input(pattern)

    patterns = [p for e in net.episodic_memory.episodes.values() for p in e['patterns']]
    if patterns:
        patterns = np.array(patterns)
        cnn_labels = [net.units.index(net.process_input(p)[0]) for p in patterns]

        min_len = min(len(patterns), len(cnn_labels))
        patterns = patterns[:min_len]
        cnn_labels = cnn_labels[:min_len]

        kmeans = KMeans(n_clusters=min(5, len(patterns)), n_init='auto').fit(patterns)
        db = DBSCAN(eps=0.4, min_samples=5).fit(patterns)

        kmeans_labels = kmeans.labels_[:min_len]
        dbscan_labels = db.labels_[:min_len]

        pcs = PCA(n_components=2).fit_transform(patterns)

        st.markdown("#### PCA View: CNN-EQIC Clusters")
        fig, ax = plt.subplots()
        ax.scatter(pcs[:, 0], pcs[:, 1], c=cnn_labels, cmap='Set2')
        ax.set_title("CNN-EQIC Neural Units as Clusters")
        st.pyplot(fig)

        st.markdown("#### PCA View: KMeans vs DBSCAN")
        fig, axs = plt.subplots(1, 2, figsize=(14, 5))
        axs[0].scatter(pcs[:, 0], pcs[:, 1], c=kmeans_labels, cmap='tab10')
        axs[0].set_title("KMeans Clustering")
        axs[1].scatter(pcs[:, 0], pcs[:, 1], c=dbscan_labels, cmap='tab10')
        axs[1].set_title("DBSCAN Clustering")
        st.pyplot(fig)

        save_path = r"C:\\Users\\oliva\\OneDrive\\Documents\\Excel doc\\CNNanalysis.xlsx"
        if st.button("Export Cluster Results to Excel"):
            save_to_excel(save_path, df, selected, cnn_labels, kmeans_labels, dbscan_labels, pcs)
            st.success(f"Exported to {save_path}")
    else:
        st.warning("No patterns extracted for clustering.")