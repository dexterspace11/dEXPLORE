# ---------------- Advanced CNN-EQIC EDA with Narrative Cluster Analysis ----------------
import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from sklearn.impute import SimpleImputer
from sklearn.preprocessing import MinMaxScaler
from sklearn.decomposition import PCA
from sklearn.cluster import KMeans
from sklearn.linear_model import LinearRegression
from sklearn.metrics import silhouette_score, davies_bouldin_score, calinski_harabasz_score
from datetime import datetime
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment

# ---------------- Memory Structures ----------------
class EpisodicMemory:
    def __init__(self):
        self.episodes = {}
        self.current_episode = None

    def create_episode(self, timestamp):
        self.current_episode = timestamp
        self.episodes[timestamp] = {'patterns': [], 'emotional_tags': []}

    def store_pattern(self, pattern, emotional_tag):
        if self.current_episode is None:
            self.create_episode(datetime.now())
        self.episodes[self.current_episode]['patterns'].append(pattern)
        self.episodes[self.current_episode]['emotional_tags'].append(emotional_tag)

class WorkingMemory:
    def __init__(self, capacity=20):
        self.capacity = capacity
        self.short_term_patterns = []
        self.temporal_context = []

    def store(self, pattern, timestamp):
        if len(self.short_term_patterns) >= self.capacity:
            self.short_term_patterns.pop(0)
            self.temporal_context.pop(0)
        self.short_term_patterns.append(pattern)
        self.temporal_context.append(timestamp)

# ---------------- Neural Unit ----------------
class HybridNeuralUnit:
    def __init__(self, position, decay_rate=100.0):
        self.position = position
        self.age = 0
        self.usage_count = 0
        self.last_spike_time = None
        self.decay_rate = decay_rate
        self.connections = []

    def distance(self, input_pattern):
        diff = np.abs(input_pattern - self.position)
        dist = np.sqrt(np.sum(diff ** 2))
        decay = np.exp(-self.age / self.decay_rate)
        return (np.exp(-2.0 * dist) + 0.5 / (1 + 0.9 * dist)) * decay

    def update_spike_time(self):
        self.last_spike_time = datetime.now()

    def communicate(self, other_units):
        for other in other_units:
            if other != self:
                dist = np.linalg.norm(self.position - other.position)
                if dist < 0.5 and other not in self.connections:
                    self.connections.append(other)

# ---------------- Neural Network ----------------
class HybridNeuralNetwork:
    def __init__(self, working_memory_capacity=20, decay_rate=100.0):
        self.units = []
        self.episodic_memory = EpisodicMemory()
        self.working_memory = WorkingMemory(capacity=working_memory_capacity)
        self.decay_rate = decay_rate

    def generate_unit(self, position):
        unit = HybridNeuralUnit(position, decay_rate=self.decay_rate)
        unit.communicate(self.units)
        self.units.append(unit)
        return unit

    def process_input(self, input_pattern):
        if not self.units:
            return self.generate_unit(input_pattern), 0.0

        similarities = [(unit, unit.distance(input_pattern)) for unit in self.units]
        similarities.sort(key=lambda x: x[1], reverse=True)
        best_unit, best_similarity = similarities[0]

        self.episodic_memory.store_pattern(input_pattern, best_similarity)
        self.working_memory.store(input_pattern, datetime.now())
        best_unit.age = 0
        best_unit.usage_count += 1
        best_unit.update_spike_time()

        if best_similarity < 0.6:
            return self.generate_unit(input_pattern), best_similarity

        return best_unit, best_similarity

# ---------------- Forecasting ----------------
def forecast_next(values, steps=1):
    model = LinearRegression()
    X = np.arange(len(values)).reshape(-1, 1)
    y = values
    model.fit(X, y)
    X_pred = np.arange(len(values), len(values) + steps).reshape(-1, 1)
    return model.predict(X_pred)

# ---------------- Narrative Generator ----------------
def interpret_level(value):
    if value < 0.33:
        return "low"
    elif value < 0.66:
        return "moderate"
    else:
        return "high"

def generate_cluster_descriptions(centroids, feature_names):
    descriptions = []
    for i, centroid in enumerate(centroids):
        traits = []
        for feat_name, val in zip(feature_names, centroid):
            level = interpret_level(val)
            traits.append(f"{level} {feat_name}")
        trait_str = ", ".join(traits)
        desc = (
            f"Cluster {i} is characterized by {trait_str}. "
            "This suggests data points in this cluster have these typical feature levels."
        )
        descriptions.append(desc)
    return descriptions

def generate_comparative_summary(centroids, feature_names):
    feature_ranges = centroids.max(axis=0) - centroids.min(axis=0)
    important_features_idx = np.where(feature_ranges > 0.2)[0]  # threshold to pick significant features

    if len(important_features_idx) == 0:
        return "Clusters show relatively similar feature profiles with minor variations."

    lines = ["Comparative summary of clusters:"]
    for idx in important_features_idx:
        feat = feature_names[idx]
        vals = centroids[:, idx]
        high_clusters = np.where(vals > 0.66)[0]
        low_clusters = np.where(vals < 0.33)[0]

        line = f"- Feature '{feat}' varies notably: "
        if len(high_clusters) > 0:
            line += f"Clusters {list(high_clusters)} show high values, "
        if len(low_clusters) > 0:
            line += f"Clusters {list(low_clusters)} show low values, "
        lines.append(line.rstrip(", ") + ".")
    return "\n".join(lines)

# ---------------- Professional Excel Export ----------------
def save_to_excel(
    save_path, 
    cluster_descriptions, 
    comparative_summary, 
    labels, 
    centroids, 
    selected_features,
    patterns
):
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Cluster Descriptions"
    for line in cluster_descriptions:
        ws1.append([line])
    ws1["A1"].font = Font(bold=True, size=14)
    ws1.column_dimensions['A'].width = 100

    ws2 = wb.create_sheet("Comparative Summary")
    for line in comparative_summary.split("\n"):
        ws2.append([line])
    ws2["A1"].font = Font(bold=True, size=14)
    ws2.column_dimensions['A'].width = 100

    ws3 = wb.create_sheet("Cluster Assignments")
    df_assign = pd.DataFrame({"Pattern Index": range(len(labels)), "Cluster": labels})
    for r in dataframe_to_rows(df_assign, index=False, header=True):
        ws3.append(r)
    for cell in ws3[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    ws3.auto_filter.ref = ws3.dimensions
    for col_cells in ws3.columns:
        length = max(len(str(cell.value)) for cell in col_cells)
        ws3.column_dimensions[col_cells[0].column_letter].width = length + 2

    ws4 = wb.create_sheet("Cluster Centroids")
    df_centroids = pd.DataFrame(centroids, columns=selected_features)
    for r in dataframe_to_rows(df_centroids, index=True, header=True):
        ws4.append(r)
    for cell in ws4[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    ws4.auto_filter.ref = ws4.dimensions
    for col_cells in ws4.columns:
        length = max(len(str(cell.value)) for cell in col_cells)
        ws4.column_dimensions[col_cells[0].column_letter].width = length + 2

    # Optional: store patterns data in a new sheet (could be large)
    ws5 = wb.create_sheet("Patterns Data")
    df_patterns = pd.DataFrame(patterns, columns=[f"F{i+1}" for i in range(patterns.shape[1])])
    for r in dataframe_to_rows(df_patterns, index=True, header=True):
        ws5.append(r)
    ws5.auto_filter.ref = ws5.dimensions
    for col_cells in ws5.columns:
        length = max(len(str(cell.value)) for cell in col_cells)
        ws5.column_dimensions[col_cells[0].column_letter].width = length + 2

    wb.save(save_path)

# ---------------- Streamlit UI ----------------
st.set_page_config(page_title="CNN-EQIC Cluster Analysis", layout="wide")
st.title("📊 CNN-EQIC: Enhanced Cluster Analysis with Narrative Explanation")

uploaded_file = st.file_uploader("Upload a CSV or Excel file", type=["csv", "xlsx"])

if uploaded_file:
    df = pd.read_csv(uploaded_file) if uploaded_file.name.endswith(".csv") else pd.read_excel(uploaded_file)
    st.markdown("### Dataset Preview")
    st.dataframe(df.head())

    numerical_cols = df.select_dtypes(include=[np.number]).columns.tolist()
    selected = st.multiselect("Select features for clustering", numerical_cols, default=numerical_cols[:4])
    window_size = st.slider("Window size (for pattern generation)", 2, 20, 5)

    if len(selected) < 2:
        st.warning("Select at least 2 numeric features for clustering.")
    else:
        clean = SimpleImputer().fit_transform(df[selected])
        scaled = MinMaxScaler().fit_transform(clean)

        net = HybridNeuralNetwork(working_memory_capacity=20, decay_rate=100.0)
        for i in range(window_size, len(scaled)):
            pattern = scaled[i - window_size:i].flatten()
            net.process_input(pattern)

        patterns = [p for e in net.episodic_memory.episodes.values() for p in e['patterns']]
        if len(patterns) == 0:
            st.warning("No patterns stored in episodic memory.")
        else:
            patterns = np.array(patterns)

            max_clusters = min(10, len(patterns))
            k = st.slider("Select number of clusters (k)", 2, max_clusters, 5)
            kmeans = KMeans(n_clusters=k, random_state=42).fit(patterns)
            labels = kmeans.labels_
            centroids = kmeans.cluster_centers_

            silhouette = silhouette_score(patterns, labels)
            db = davies_bouldin_score(patterns, labels)
            ch = calinski_harabasz_score(patterns, labels)

            st.markdown("### Clustering Evaluation Metrics")
            st.metric("Silhouette Score", f"{silhouette:.3f}")
            st.metric("Davies-Bouldin Index", f"{db:.3f}")
            st.metric("Calinski-Harabasz Score", f"{ch:.1f}")

            pca = PCA(n_components=2)
            pcs = pca.fit_transform(patterns)
            df_pca = pd.DataFrame(pcs, columns=['PC1', 'PC2'])
            df_pca['Cluster'] = labels

            fig, ax = plt.subplots()
            sns.scatterplot(x='PC1', y='PC2', hue='Cluster', data=df_pca, palette='tab10', ax=ax)
            ax.set_title("PCA Cluster Projection")
            st.pyplot(fig)

            # Generate cluster narratives
            cluster_descriptions = generate_cluster_descriptions(centroids, selected)
            st.markdown("### Cluster Descriptions")
            for desc in cluster_descriptions:
                st.write(desc)

            # Comparative summary of clusters
            comp_summary = generate_comparative_summary(centroids, selected)
            st.markdown("### Comparative Summary of Clusters")
            st.text(comp_summary)

            export_path = r"C:\Users\oliva\OneDrive\Documents\Excel doc\DNNanalysis.xlsx"
            if st.button("Export Analysis Results to Excel"):
                save_to_excel(
                    save_path=export_path,
                    cluster_descriptions=cluster_descriptions,
                    comparative_summary=comp_summary,
                    labels=labels,
                    centroids=centroids,
                    selected_features=selected,
                    patterns=patterns
                )
                st.success(f"Analysis exported successfully to:\n{export_path}")
