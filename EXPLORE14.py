# ---------------- Advanced CNN-EQIC EDA with Professional Excel Export ----------------
import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from sklearn.impute import SimpleImputer
from sklearn.preprocessing import MinMaxScaler
from sklearn.decomposition import PCA
from sklearn.feature_selection import mutual_info_regression
from scipy.stats import zscore
from sklearn.cluster import KMeans, DBSCAN
from sklearn.linear_model import LinearRegression
from datetime import datetime
import os
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment

# ---------------- Memory Structures ----------------
class EpisodicMemory:
    def __init__(self):
        self.episodes = {}
        self.current_episode = None

    def create_episode(self, timestamp):
        self.current_episode = timestamp
        self.episodes[timestamp] = {'patterns': [], 'emotional_tags': []}

    def store_pattern(self, pattern, emotional_tag):
        if self.current_episode is None:
            self.create_episode(datetime.now())
        self.episodes[self.current_episode]['patterns'].append(pattern)
        self.episodes[self.current_episode]['emotional_tags'].append(emotional_tag)

class WorkingMemory:
    def __init__(self, capacity=20):
        self.capacity = capacity
        self.short_term_patterns = []
        self.temporal_context = []

    def store(self, pattern, timestamp):
        if len(self.short_term_patterns) >= self.capacity:
            self.short_term_patterns.pop(0)
            self.temporal_context.pop(0)
        self.short_term_patterns.append(pattern)
        self.temporal_context.append(timestamp)

# ---------------- Neural Unit ----------------
class HybridNeuralUnit:
    def __init__(self, position, decay_rate=100.0):
        self.position = position
        self.age = 0
        self.usage_count = 0
        self.last_spike_time = None
        self.decay_rate = decay_rate
        self.connections = []  # Inter-unit links

    def distance(self, input_pattern):
        diff = np.abs(input_pattern - self.position)
        dist = np.sqrt(np.sum(diff ** 2))
        decay = np.exp(-self.age / self.decay_rate)
        return (np.exp(-2.0 * dist) + 0.5 / (1 + 0.9 * dist)) * decay

    def update_spike_time(self):
        self.last_spike_time = datetime.now()

    def communicate(self, other_units):
        for other in other_units:
            if other != self:
                dist = np.linalg.norm(self.position - other.position)
                if dist < 0.5 and other not in self.connections:
                    self.connections.append(other)

# ---------------- Neural Network ----------------
class HybridNeuralNetwork:
    def __init__(self, working_memory_capacity=20, decay_rate=100.0):
        self.units = []
        self.episodic_memory = EpisodicMemory()
        self.working_memory = WorkingMemory(capacity=working_memory_capacity)
        self.decay_rate = decay_rate

    def generate_unit(self, position):
        unit = HybridNeuralUnit(position, decay_rate=self.decay_rate)
        unit.communicate(self.units)
        self.units.append(unit)
        return unit

    def process_input(self, input_pattern):
        if not self.units:
            return self.generate_unit(input_pattern), 0.0

        similarities = [(unit, unit.distance(input_pattern)) for unit in self.units]
        similarities.sort(key=lambda x: x[1], reverse=True)
        best_unit, best_similarity = similarities[0]

        self.episodic_memory.store_pattern(input_pattern, best_similarity)
        self.working_memory.store(input_pattern, datetime.now())
        best_unit.age = 0
        best_unit.usage_count += 1
        best_unit.update_spike_time()

        if best_similarity < 0.6:
            return self.generate_unit(input_pattern), best_similarity

        return best_unit, best_similarity

# ---------------- Forecasting ----------------
def forecast_next(values, steps=1):
    model = LinearRegression()
    X = np.arange(len(values)).reshape(-1, 1)
    y = values
    model.fit(X, y)
    X_pred = np.arange(len(values), len(values) + steps).reshape(-1, 1)
    return model.predict(X_pred)

# ---------------- Professional Excel Export ----------------
def save_to_excel_professional(save_path, df, selected, timestamps, similarities, kmeans_labels, dbscan_labels, pcs, outliers, usage_counts, best_params, best_score):
    wb = openpyxl.Workbook()

    # Sheet 1: Clean Data
    ws1 = wb.active
    ws1.title = "Clean Data"
    for r in dataframe_to_rows(df[selected], index=False, header=True):
        ws1.append(r)
    for cell in ws1[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    ws1.auto_filter.ref = ws1.dimensions
    for col_cells in ws1.columns:
        length = max(len(str(cell.value)) for cell in col_cells)
        ws1.column_dimensions[col_cells[0].column_letter].width = length + 2

    # Sheet 2: Pattern Similarities Over Time
    ws2 = wb.create_sheet("Pattern Similarities")
    sim_df = pd.DataFrame({"Timestamp_Index": timestamps, "Pattern Similarity": similarities})
    for r in dataframe_to_rows(sim_df, index=False, header=True):
        ws2.append(r)
    for cell in ws2[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    ws2.auto_filter.ref = ws2.dimensions
    for col_cells in ws2.columns:
        length = max(len(str(cell.value)) for cell in col_cells)
        ws2.column_dimensions[col_cells[0].column_letter].width = length + 2

    # Sheet 3: Clustering Labels
    ws3 = wb.create_sheet("Cluster Labels")
    cluster_df = pd.DataFrame({
        "Pattern Index": range(len(kmeans_labels)),
        "KMeans Cluster Label": kmeans_labels,
        "DBSCAN Cluster Label": dbscan_labels
    })
    for r in dataframe_to_rows(cluster_df, index=False, header=True):
        ws3.append(r)
    for cell in ws3[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    ws3.auto_filter.ref = ws3.dimensions
    for col_cells in ws3.columns:
        length = max(len(str(cell.value)) for cell in col_cells)
        ws3.column_dimensions[col_cells[0].column_letter].width = length + 2

    # Sheet 4: PCA Components
    ws4 = wb.create_sheet("PCA Components")
    pcs_df = pd.DataFrame(pcs, columns=['PC1', 'PC2'])
    for r in dataframe_to_rows(pcs_df, index=False, header=True):
        ws4.append(r)
    for cell in ws4[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    ws4.auto_filter.ref = ws4.dimensions
    for col_cells in ws4.columns:
        length = max(len(str(cell.value)) for cell in col_cells)
        ws4.column_dimensions[col_cells[0].column_letter].width = length + 2

    # Sheet 5: Outliers
    ws5 = wb.create_sheet("Outliers (Z-score > 3)")
    if outliers.empty:
        ws5.append(["No outliers detected using Z-score method."])
    else:
        for r in dataframe_to_rows(outliers, index=False, header=True):
            ws5.append(r)
        for cell in ws5[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        ws5.auto_filter.ref = ws5.dimensions
        for col_cells in ws5.columns:
            length = max(len(str(cell.value)) for cell in col_cells)
            ws5.column_dimensions[col_cells[0].column_letter].width = length + 2

    # Sheet 6: Neural Unit Usage Counts
    ws6 = wb.create_sheet("Neural Unit Usage")
    usage_df = pd.DataFrame(usage_counts, columns=['Usage Count'])
    for r in dataframe_to_rows(usage_df, index=True, header=True):
        ws6.append(r)
    for cell in ws6[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    ws6.auto_filter.ref = ws6.dimensions
    for col_cells in ws6.columns:
        length = max(len(str(cell.value)) for cell in col_cells)
        ws6.column_dimensions[col_cells[0].column_letter].width = length + 2

    # Sheet 7: Summary & Interpretation
    ws7 = wb.create_sheet("Summary & Interpretation")
    summary_lines = [
        "CNN-EQIC Analysis Summary",
        "",
        "Best Parameters:",
        f"  - Working Memory Capacity: {best_params.get('working_memory_capacity', 'N/A')}",
        f"  - Decay Rate: {best_params.get('decay_rate', 'N/A')}",
        "",
        f"Best Similarity Score: {best_score:.4f}",
        f"Number of Neural Units Generated: {len(usage_counts)}",
        "",
        "Clustering:",
        f"  - KMeans Clusters: {len(np.unique(kmeans_labels)) if kmeans_labels is not None else 'N/A'}",
        f"  - DBSCAN Outliers Detected: {np.sum(dbscan_labels == -1) if dbscan_labels is not None else 'N/A'}",
        "",
        f"Outliers Detected by Z-score Method: {len(outliers)}",
        "",
        "Notes:",
        "- Higher neural unit usage counts suggest more frequent pattern matches.",
        "- Clusters group similar patterns together; outliers may represent anomalies or noise.",
        "- PCA projection helps visualize cluster separation in 2D space.",
        "- Forecasting provides predicted similarity trends for short-term patterns.",
        "",
        "Please refer to individual sheets for detailed data."
    ]
    for line in summary_lines:
        ws7.append([line])
    # Bold the title
    ws7["A1"].font = Font(bold=True, size=14)
    ws7.column_dimensions['A'].width = 100
    ws7.sheet_view.showGridLines = False

    # Save workbook
    wb.save(save_path)

# ---------------- Streamlit UI ----------------
st.set_page_config(page_title="Advanced CNN-EQIC EDA", layout="wide")
st.title("📊 CNN-EQIC: Dynamic Clustering + Forecasting + Anomaly Detection")

uploaded_file = st.file_uploader("Upload a CSV or Excel file", type=["csv", "xlsx"])

if uploaded_file:
    df = pd.read_csv(uploaded_file) if uploaded_file.name.endswith(".csv") else pd.read_excel(uploaded_file)
    st.markdown("### Preview")
    st.dataframe(df.head())

    numerical_cols = df.select_dtypes(include=[np.number]).columns.tolist()
    selected = st.multiselect("Select features", numerical_cols, default=numerical_cols[:4])
    window_size = st.slider("Window Size", 2, 20, 5)

    clean = SimpleImputer().fit_transform(df[selected])
    scaled = MinMaxScaler().fit_transform(clean)

    st.markdown("### Tuning & Dynamic Network Clustering")
    param_grid = [
        {'working_memory_capacity': c, 'decay_rate': d}
        for c in [10, 20] for d in [50.0, 100.0]
    ]
    def evaluate(net, data):
        scores = []
        for i in range(window_size, len(data)):
            pattern = data[i - window_size:i].flatten()
            _, sim = net.process_input(pattern)
            scores.append(sim)
        return np.mean(scores)

    best_score = -np.inf
    best_params = {}
    for params in param_grid:
        net = HybridNeuralNetwork(**params)
        score = evaluate(net, scaled)
        if score > best_score:
            best_params, best_score = params, score

    st.success(f"Best Params: {best_params}, Best Similarity: {best_score:.4f}")

    net = HybridNeuralNetwork(**best_params)
    similarities, timestamps = [], []
    for i in range(window_size, len(scaled)):
        pattern = scaled[i - window_size:i].flatten()
        _, sim = net.process_input(pattern)
        similarities.append(sim)
        timestamps.append(i)

    st.line_chart(pd.DataFrame({"Similarity": similarities}, index=timestamps))
    st.markdown("### CNN Unit Usage Frequency")
    usage_counts = [u.usage_count for u in net.units]
    st.bar_chart(usage_counts)

    st.markdown("### Clustering & Forecasting")
    patterns = [p for e in net.episodic_memory.episodes.values() for p in e['patterns']]
    if patterns:
        patterns = np.array(patterns)
        kmeans = KMeans(n_clusters=min(5, len(patterns))).fit(patterns)
        st.dataframe(pd.DataFrame({'Cluster': kmeans.labels_}))

        st.markdown("#### Forecasting Next 3 Steps")
        forecast_values = forecast_next(similarities[-10:], 3)
        st.write(f"Forecast: {forecast_values}")

        st.markdown("#### DBSCAN Anomaly Detection")
        db = DBSCAN(eps=0.4, min_samples=5).fit(patterns)
        outliers_idx = np.where(db.labels_ == -1)[0]
        st.write(f"Outliers Detected: {len(outliers_idx)}")
        st.dataframe(pd.DataFrame(patterns[outliers_idx], columns=[f"F{i}" for i in range(patterns.shape[1])]))

        st.markdown("#### PCA View")
        pcs = PCA(n_components=2).fit_transform(patterns)
        fig, ax = plt.subplots()
        scatter = ax.scatter(pcs[:, 0], pcs[:, 1], c=kmeans.labels_, cmap='Set1')
        ax.set_title("Clustering Centroid Projection")
        st.pyplot(fig)

        st.markdown("#### Correlation Matrix")
        fig, ax = plt.subplots()
        sns.heatmap(pd.DataFrame(clean, columns=selected).corr(), annot=True, cmap="coolwarm", ax=ax)
        st.pyplot(fig)

        st.markdown("#### Z-score Outlier Detection")
        zs = np.abs(zscore(clean))
        outliers_df = pd.DataFrame(clean, columns=selected)[(zs > 3).any(axis=1)]
        st.write(f"Outliers: {len(outliers_df)}")
        st.dataframe(outliers_df)

        # Export to Excel button
        save_path = r"C:\Users\oliva\OneDrive\Documents\Excel doc\CNNanalysis.xlsx"
        if st.button("Export Results to Excel"):
            save_to_excel_professional(
                save_path=save_path,
                df=df,
                selected=selected,
                timestamps=timestamps,
                similarities=similarities,
                kmeans_labels=kmeans.labels_,
                dbscan_labels=db.labels_,
                pcs=pcs,
                outliers=outliers_df,
                usage_counts=usage_counts,
                best_params=best_params,
                best_score=best_score
            )
            st.success(f"Results exported to {save_path}")
    else:
        st.warning("No patterns stored in memory.")
