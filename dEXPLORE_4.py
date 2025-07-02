# ---------------- Advanced CNN-EQIC EDA with Narrative Cluster Analysis ----------------
import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from sklearn.impute import SimpleImputer
from sklearn.preprocessing import MinMaxScaler
from sklearn.decomposition import PCA
from sklearn.cluster import KMeans
from sklearn.metrics import silhouette_score, davies_bouldin_score, calinski_harabasz_score
from scipy.cluster.hierarchy import linkage, dendrogram
from datetime import datetime
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment
import joblib
import io
import base64

# ---------------- Memory Structures ----------------
class EpisodicMemory:
    def __init__(self):
        self.episodes = {}
        self.current_episode = None

    def create_episode(self, timestamp):
        self.current_episode = timestamp
        self.episodes[timestamp] = {'patterns': [], 'emotional_tags': []}

    def store_pattern(self, pattern, emotional_tag):
        if self.current_episode is None:
            self.create_episode(datetime.now())
        self.episodes[self.current_episode]['patterns'].append(pattern)
        self.episodes[self.current_episode]['emotional_tags'].append(emotional_tag)

class WorkingMemory:
    def __init__(self, capacity=20):
        self.capacity = capacity
        self.short_term_patterns = []
        self.temporal_context = []

    def store(self, pattern, timestamp):
        if len(self.short_term_patterns) >= self.capacity:
            self.short_term_patterns.pop(0)
            self.temporal_context.pop(0)
        self.short_term_patterns.append(pattern)
        self.temporal_context.append(timestamp)

# ---------------- Neural Unit ----------------
class HybridNeuralUnit:
    def __init__(self, position, decay_rate=100.0):
        self.position = position
        self.age = 0
        self.usage_count = 0
        self.last_spike_time = None
        self.decay_rate = decay_rate
        self.connections = []

    def distance(self, input_pattern):
        diff = np.abs(input_pattern - self.position)
        dist = np.sqrt(np.sum(diff ** 2))
        decay = np.exp(-self.age / self.decay_rate)
        return (np.exp(-2.0 * dist) + 0.5 / (1 + 0.9 * dist)) * decay

    def update_spike_time(self):
        self.last_spike_time = datetime.now()

    def communicate(self, other_units):
        for other in other_units:
            if other != self:
                dist = np.linalg.norm(self.position - other.position)
                if dist < 0.5 and other not in self.connections:
                    self.connections.append(other)

# ---------------- Neural Network ----------------
class HybridNeuralNetwork:
    def __init__(self, working_memory_capacity=20, decay_rate=100.0):
        self.units = []
        self.episodic_memory = EpisodicMemory()
        self.working_memory = WorkingMemory(capacity=working_memory_capacity)
        self.decay_rate = decay_rate

    def generate_unit(self, position):
        unit = HybridNeuralUnit(position, decay_rate=self.decay_rate)
        unit.communicate(self.units)
        self.units.append(unit)
        return unit

    def process_input(self, input_pattern):
        if not self.units:
            return self.generate_unit(input_pattern), 0.0

        similarities = [(unit, unit.distance(input_pattern)) for unit in self.units]
        similarities.sort(key=lambda x: x[1], reverse=True)
        best_unit, best_similarity = similarities[0]

        self.episodic_memory.store_pattern(input_pattern, best_similarity)
        self.working_memory.store(input_pattern, datetime.now())
        best_unit.age = 0
        best_unit.usage_count += 1
        best_unit.update_spike_time()

        if best_similarity < 0.6:
            return self.generate_unit(input_pattern), best_similarity

        return best_unit, best_similarity

# ---------------- Narrative Generator ----------------
def interpret_level(value):
    if value < 0.33:
        return "low"
    elif value < 0.66:
        return "moderate"
    else:
        return "high"

def generate_cluster_descriptions(centroids, feature_names):
    descriptions = []
    for i, centroid in enumerate(centroids):
        traits = []
        for feat_name, val in zip(feature_names, centroid):
            level = interpret_level(val)
            traits.append(f"{level} {feat_name}")
        trait_str = ", ".join(traits)
        desc = f"Cluster {i} is characterized by {trait_str}."
        descriptions.append(desc)
    return descriptions

def generate_comparative_summary(centroids, feature_names):
    feature_ranges = centroids.max(axis=0) - centroids.min(axis=0)
    important_features_idx = np.where(feature_ranges > 0.2)[0]
    if len(important_features_idx) == 0:
        return "Clusters show similar profiles with minimal variation."

    lines = ["Comparative summary of cluster differences:"]
    for idx in important_features_idx:
        feat = feature_names[idx] if idx < len(feature_names) else f"Feature_{idx}"
        vals = centroids[:, idx]
        high_clusters = np.where(vals > 0.66)[0]
        low_clusters = np.where(vals < 0.33)[0]
        line = f"- '{feat}' is high in clusters {list(high_clusters)} and low in clusters {list(low_clusters)}"
        lines.append(line)
    return "\n".join(lines)

# ---------------- Streamlit App ----------------
st.set_page_config(page_title="CNN-EQIC Cluster Analysis", layout="wide")
st.title("📊 CNN-EQIC: Dynamic Clustering with Explanations")

uploaded_file = st.file_uploader("Upload a CSV or Excel file", type=["csv", "xlsx"])
if uploaded_file:
    df = pd.read_csv(uploaded_file) if uploaded_file.name.endswith(".csv") else pd.read_excel(uploaded_file)
    st.dataframe(df.head())

    numerical_cols = df.select_dtypes(include=[np.number]).columns.tolist()
    default_selection = numerical_cols[:min(4, len(numerical_cols))]
    selected = st.multiselect("Select features", numerical_cols, default=default_selection)
    window_size = st.slider("Window size", 2, 20, 5)

    if len(selected) >= 2:
        clean = SimpleImputer().fit_transform(df[selected])
        scaler = MinMaxScaler().fit(clean)  # Fit here so we can save scaler later
        scaled = scaler.transform(clean)

        net = HybridNeuralNetwork(working_memory_capacity=20, decay_rate=100.0)
        for i in range(window_size, len(scaled)):
            pattern = scaled[i - window_size:i].flatten()
            net.process_input(pattern)

        patterns = [p for e in net.episodic_memory.episodes.values() for p in e['patterns']]
        if len(patterns) > 0:
            patterns = np.array(patterns)

            expanded_feature_names = [f"{feat}_t{t}" for t in range(window_size) for feat in selected]

            k = st.slider("Number of clusters", 2, min(10, len(patterns)), 5)
            kmeans = KMeans(n_clusters=k, random_state=42).fit(patterns)
            labels = kmeans.labels_
            centroids = kmeans.cluster_centers_

            silhouette = silhouette_score(patterns, labels)
            db = davies_bouldin_score(patterns, labels)
            ch = calinski_harabasz_score(patterns, labels)
            st.metric("Silhouette Score", f"{silhouette:.3f}")
            st.metric("Davies-Bouldin Index", f"{db:.3f}")
            st.metric("Calinski-Harabasz Score", f"{ch:.1f}")

            st.subheader("📌 PCA Cluster Projection")
            pca = PCA(n_components=2).fit_transform(patterns)
            pca_df = pd.DataFrame(pca, columns=['PC1', 'PC2'])
            pca_df['Cluster'] = labels
            fig, ax = plt.subplots()
            sns.scatterplot(data=pca_df, x='PC1', y='PC2', hue='Cluster', palette='tab10', ax=ax)
            st.pyplot(fig)

            st.subheader("🧠 Cluster Interpretations")
            cluster_descriptions = generate_cluster_descriptions(centroids, expanded_feature_names)
            for desc in cluster_descriptions:
                st.markdown(f"- {desc}")

            st.subheader("🧮 Comparative Feature Summary")
            comp_summary = generate_comparative_summary(centroids, expanded_feature_names)
            st.text(comp_summary)

            st.subheader("📊 Correlation Matrix")
            fig2, ax2 = plt.subplots()
            sns.heatmap(pd.DataFrame(clean, columns=selected).corr(), annot=True, cmap="coolwarm", ax=ax2)
            st.pyplot(fig2)

            st.subheader("🧬 Dendrogram (Hierarchical Clustering)")
            linkage_matrix = linkage(patterns, method='ward')
            fig3, ax3 = plt.subplots(figsize=(10, 4))
            dendrogram(linkage_matrix, truncate_mode='level', p=5, ax=ax3)
            st.pyplot(fig3)

            st.subheader("📘 Academic Report Summary")
            st.markdown("""
                This clustering analysis segmented the data into meaningful groups based on feature patterns over time.
                Each cluster reveals typical configurations, such as consistently high or low values in specific features
                across time windows. Clusters with strong separation (high silhouette scores) indicate clear behavioral
                differentiation. The dendrogram and PCA plots help visualize these relationships.
            """)

            if st.button("Export to Excel"):
                export_path = r"C:\\Users\\oliva\\OneDrive\\Documents\\Excel doc\\DNNanalysis.xlsx"
                wb = openpyxl.Workbook()
                ws1 = wb.active
                ws1.title = "Cluster Descriptions"
                for line in cluster_descriptions:
                    ws1.append([line])
                ws2 = wb.create_sheet("Comparative Summary")
                for line in comp_summary.split("\n"):
                    ws2.append([line])
                ws3 = wb.create_sheet("Cluster Labels")
                df_assign = pd.DataFrame({"Pattern Index": range(len(labels)), "Cluster": labels})
                for r in dataframe_to_rows(df_assign, index=False, header=True):
                    ws3.append(r)
                ws4 = wb.create_sheet("Centroids")
                df_centroids = pd.DataFrame(centroids, columns=expanded_feature_names)
                for r in dataframe_to_rows(df_centroids, index=False, header=True):
                    ws4.append(r)
                # Add raw data by cluster
                for i in range(k):
                    cluster_data = patterns[labels == i]
                    df_cluster = pd.DataFrame(cluster_data, columns=expanded_feature_names)
                    ws = wb.create_sheet(f"Raw Cluster {i}")
                    for r in dataframe_to_rows(df_cluster, index=False, header=True):
                        ws.append(r)
                # Save scaler as base64 string in a new sheet "Scaler"
                scaler_bytes_io = io.BytesIO()
                joblib.dump(scaler, scaler_bytes_io)
                scaler_bytes = scaler_bytes_io.getvalue()
                scaler_b64 = base64.b64encode(scaler_bytes).decode('utf-8')
                ws_scaler = wb.create_sheet("Scaler")
                # Write in chunks of 1000 chars for readability
                chunk_size = 1000
                for i in range(0, len(scaler_b64), chunk_size):
                    ws_scaler.append([scaler_b64[i:i+chunk_size]])
                # Auto-adjust column widths
                for ws in wb.worksheets:
                    for col_cells in ws.columns:
                        length = max(len(str(cell.value) or "") for cell in col_cells)
                        ws.column_dimensions[col_cells[0].column_letter].width = length + 2
                wb.save(export_path)
                st.success(f"Exported results and scaler to {export_path}")
        else:
            st.warning("No patterns were stored. Try selecting more diverse features or reducing window size.")
